"""
Generate the Worship Wheel "Product Launch" workbook.

A simple collaboration sheet for Charl / Schulte to populate the fields for each
new product CTA card before it's entered into the admin. Two tabs:

  1. "Merge Fields" — the available {tokens} and a linguistic example of each.
  2. "Products"     — one row per product; columns mirror the admin form fields.

Run:  python3 specs/009-product-cta-cards/build-product-sheet.py
Out:  specs/009-product-cta-cards/Worship-Wheel-Product-Launch-Sheet.xlsx

The .xlsx is intended to be uploaded to Google Drive for collaboration, so this
keeps styling deliberately minimal: bold + light-gray header row, sensible
column widths, wrapped text on the long copy columns.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="E8E8E7")  # light gray (brand neutral-200)
HEADER_FONT = Font(bold=True, color="100F0E")
NOTE_FONT = Font(italic=True, color="615F5B")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Tab 1 — Merge Fields
# ---------------------------------------------------------------------------
MERGE_FIELDS = [
    ("{firstName}", "Great work, {firstName} — here's exactly where to focus next."),
    ("{overallScore}", "You scored {overallScore} on the Worship Wheel assessment."),
    ("{overallScore}/80", "Based on your score of {overallScore}/80, you're closer than you think."),
    ("{archetypeName}", "As a {archetypeName}, your fastest win is just one habit away."),
    ("{weakestElement}", "Right now your biggest opportunity is your {weakestElement}."),
]

MERGE_NOTES = [
    "Tip: there's no token that includes the \"/80\" — just type it yourself, e.g. {overallScore}/80, since the max is always 80.",
    "Merge fields work in: Headline, Sub-headline, CTA headline and CTA copy.",
    "If a value is missing (or you mistype a token), it's simply removed — a visitor never sees a raw {token}, and spacing/punctuation is tidied up automatically.",
]


def build_merge_tab(wb):
    ws = wb.active
    ws.title = "Merge Fields"
    ws.append(["Merge field", "Example usage in a sentence"])
    for token, example in MERGE_FIELDS:
        ws.append([token, example])

    style_header(ws, 2)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70
    for row in ws.iter_rows(min_row=2, max_row=1 + len(MERGE_FIELDS)):
        row[0].alignment = TOP
        row[1].alignment = WRAP_TOP

    # Notes a couple of rows below the table.
    note_row = len(MERGE_FIELDS) + 4
    for i, note in enumerate(MERGE_NOTES):
        cell = ws.cell(row=note_row + i, column=1, value=note)
        cell.font = NOTE_FONT
        cell.alignment = WRAP_TOP


# ---------------------------------------------------------------------------
# Tab 2 — Products
# ---------------------------------------------------------------------------
# (header, column width, wrap?)  — order mirrors the admin "New product" form.
COLUMNS = [
    ("Internal name (admin only)", 28, False),
    ("Code (optional — leave blank to auto-generate)", 24, False),
    ("Status (draft / active)", 18, False),
    ("Headline  ⟨merge fields⟩", 32, True),
    ("Sub-headline (optional)  ⟨merge fields⟩", 36, True),
    ("Video URL (optional — Vimeo)", 30, False),
    ("Eyebrow", 24, False),
    ("CTA headline  ⟨merge fields⟩", 32, True),
    ("CTA copy  ⟨merge fields⟩", 46, True),
    ("Button label", 22, False),
    ("Button link", 32, False),
]

# One example row so the format is obvious — clearly marked for deletion.
EXAMPLE_ROW = [
    "EXAMPLE — delete this row",
    "src",
    "draft",
    "The Uneven Intermediate",
    "You've got real strengths — and a couple of gaps holding the whole thing back.",
    "https://vimeo.com/123456789",
    "READY TO LEVEL UP?",
    "Start the 90-Day Challenge",
    "Based on your score of {overallScore}/80, {firstName}, your fastest win is fixing your {weakestElement}.",
    "Start the Challenge",
    "https://worshipguitarskills.com/challenge",
]


def build_products_tab(wb):
    ws = wb.create_sheet("Products")
    ws.append([c[0] for c in COLUMNS])
    ws.append(EXAMPLE_ROW)
    # A few blank rows ready to fill.
    for _ in range(20):
        ws.append([None] * len(COLUMNS))

    style_header(ws, len(COLUMNS))
    for i, (_, width, wrap) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    align = WRAP_TOP
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(COLUMNS)):
        for i, cell in enumerate(row):
            cell.alignment = WRAP_TOP if COLUMNS[i][2] else TOP

    # Grey out the example row so it reads as a sample, not real data.
    for cell in ws[2]:
        cell.font = NOTE_FONT


def main():
    wb = Workbook()
    build_merge_tab(wb)
    build_products_tab(wb)
    out = Path(__file__).with_name("Worship-Wheel-Product-Launch-Sheet.xlsx")
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
